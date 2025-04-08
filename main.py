from tkinter import *
from tkinter import filedialog
import os
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Border, Side

root = Tk()
root.title('Sasol De Bron - GP Analyzer ')
root.geometry('230x150')

def gp_analysis():
    # Get Less than amounts
    less = int(less_than_input.get())
    less_than_input.delete(0,END)

    # Select .csv file
    path = os.getcwd()
    file = filedialog.askopenfilename(initialdir=path, title='Select A File', filetypes=(("CSV Files", '*.csv'),))
    
    # Read in data from selected csv file
    read_info = pd.read_csv(file, header=1).dropna()
    data = read_info.to_numpy().tolist()

    # Remove unwated data from data list
    simple_list = []
    for x in data:
        barcode = x[0].strip()
        name = x[1].strip()
        amt = x[2].strip()

        simple_list.append([barcode, name, amt, x[3], x[4], x[5], x[6], x[7], x[8], x[9]])

    # Run through list and execute less than amount
    printing_list = []
    for x in simple_list:
        if x[-2] <= less:
            printing_list.append(x)

    items = len(printing_list) + 2

    # Create Wwork book   
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock GP List'

    # Set Column Width 
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35

    # Set column names
    ws['C1'] = 'Pack'
    ws['E1'] = 'Desired %'
    ws['F1'] = 'Sugg Incl'
    ws['G1'] = 'Selling A'
    ws['H1'] = 'Var Amt on'
    ws['I1'] = 'GP % on'
    ws['J1'] = 'Var % on'

    ws['A2'] = 'Pack Code'
    ws['B2'] = 'Pack Description'
    ws['C2'] = 'Size'
    ws['D2'] = 'Sys Cost'
    ws['E2'] = 'on Sell A'
    ws['F2'] = 'Selling A'
    ws['G2'] = 'Inclusive'
    ws['H2'] = 'Sell A Incl'
    ws['I2'] = 'Sys Cost'
    ws['J2'] = 'Sys Cost'

    # Loop through sheet totals
    i = 3
    while i <= items:
        for stock in printing_list:
            ws[f'A{i}'] = stock[0]
            ws[f'B{i}'] = stock[1]
            ws[f'C{i}'] = stock[2]
            ws[f'D{i}'] = stock[3]
            ws[f'E{i}'] = stock[4]
            ws[f'F{i}'] = stock[5]
            ws[f'G{i}'] = stock[6]
            ws[f'H{i}'] = stock[7]
            ws[f'I{i}'] = stock[8]
            ws[f'J{i}'] = stock[9]
            
            ws[f'A{i}'].number_format = '0'
            i += 1

    info.config(text='Analysis Complete')
    f_name = file_name_input.get()
    if f_name == '':
        wb.save(f'GP Analysis Less than {less}.xlsx')
    else:
        wb.save(f'{f_name}.xlsx')
        file_name_input.delete(0,END)

# Less than widget
less_than = Label(root, text='Less Than')
less_than.grid(row=0, column=0, padx=(10,0), pady=(10,0))

less_than_input = Entry(root, text='Less Than')
less_than_input.grid(row=0, column=1, padx=(10,0), pady=(10,0))

# File name widget
file_name = Label(root, text='File Name')
file_name.grid(row=2, column=0, padx=(10,0), pady=(10,0))

file_name_input = Entry(root, text='File Name')
file_name_input.grid(row=2, column=1, padx=(10,0), pady=(10,0))

start_analysis_btn = Button(root, text='Start Analysis', command=gp_analysis)
start_analysis_btn.grid(row=4, column=0, columnspan=2, sticky=W+E, padx=(10,10), pady=(10,10))

info = Label(root, text='')
info.grid(row=5, column=0, columnspan=2)

# Run program
root.mainloop()



